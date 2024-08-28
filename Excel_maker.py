import openpyxl
from openpyxl.workbook import Workbook
from openpyxl import load_workbook

from openpyxl.worksheet.table import Table, TableStyleInfo

from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

from openpyxl.styles import Font, Alignment, numbers

import pandas as pd
import chardet

file_path = "data/eo_2024_check.csv"

with open(file_path, 'rb') as file:
    raw_data = file.read()
    result = chardet.detect(raw_data)
    detected_encoding = result['encoding']

print(f"Detected encoding: {detected_encoding}")

df = pd.read_csv(file_path, sep=',', encoding=detected_encoding)
print(df.head(1))

wb = Workbook()
ws = wb.active

# SustainablePop_column = ws['C'] # No because this makes a Tuple

for row in dataframe_to_rows(df, index=False, header=True):
   ws.append(row)

default_data_width = 16

column_widths = {'B': 24}
for col in range(ord('C'), ord('H')+ 1):
    column_widths[chr(col)] = default_data_width

for col, width in column_widths.items():
    ws.column_dimensions[col].width = width

# column_widths = {'B': 20, 'C': 12, 'D': 12, 'E': 12, 'F': 12, 'G': 12, 'H': 12}
# for col, width in column_widths.items():
#     ws.column_dimensions[col].width = width

# remove decimal places for columns C to H
for col in range(3, 9): 
    for cell in ws[get_column_letter(col)][1:]:  # Start from row 1 to skip header
        if cell.value is not None and isinstance(cell.value, (int, float)):
            cell.value = round(cell.value, 0)

# for cell in ws['C'][1:]:  # start from row 1 to skip header
#    cell.value = round(cell.value, 0)
   #  cell.value = cell.value / 100
   #  cell.number_format = numbers.FORMAT_PERCENTAGE

# ws.column_dimensions['B'].width = 20
# ws.column_dimensions['C'].width = 12
# ws.column_dimensions['D'].width = 12

style = TableStyleInfo(name="Tab1eStyleMedium9", showRowStripes=True)
table = Table(displayName="Percentage_difference",
              ref="A1:" + get_column_letter(ws.max_column) + str(ws.max_row))
table.tableStyleInfo = style

ws.insert_rows(idx=1)
title = ws['A1']
title.value = "Percentage difference between 2024 and 2023 data to prompt spot-checks"

ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)

ws.row_dimensions[1].height = 24
ws.row_dimensions[2].height = 28

for row in range(3, ws.max_row + 1):
    ws.row_dimensions[row].height = 18

for row in ws['A2':'H2']:
   for cell in row:
      cell.alignment = Alignment(vertical='center')
      cell.font = Font(bold=True)

def suspop_highlight(val):
   color = '#EE9090' if val > 5 else ''
   return 'background-color: {}'.format(color)

# ws.style.applymap(suspop_highlight, subset=pd.IndexSlice[:,['Maximum_Pop_pct_diff']])

from openpyxl.styles import PatternFill
# fill pattern for highlighted cells
highlight_fill = PatternFill(start_color="EE9090", end_color="EE9090", fill_type="solid")

col_letter = 'C' # Maximum sustainable population
for cell in ws[col_letter][3:]:  # skip title & header by beginning at row 3
    if cell.value is not None and isinstance(cell.value, (int, float)):
        if cell.value >= 10 or cell.value <= -10:
            cell.fill = highlight_fill

col_letter = 'E' # Population total
for cell in ws[col_letter][3:]:
    if cell.value is not None and isinstance(cell.value, (int, float)):
        if cell.value >= 4 or cell.value <= -3:
            cell.fill = highlight_fill

col_letter = 'G' # Species threatened
for cell in ws[col_letter][3:]:
    if cell.value is not None and isinstance(cell.value, (int, float)):
        if cell.value >= 10 or cell.value <= -10:
            cell.fill = highlight_fill

col_letter = 'H' # GDP per capita
for cell in ws[col_letter][3:]:
    if cell.value is not None and isinstance(cell.value, (int, float)):
        if cell.value >= 20 or cell.value <= -20:
            cell.fill = highlight_fill

wb.save('data/checking2.xlsx')
